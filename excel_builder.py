"""
excel_builder.py
-----------------
Writes the SSL pipeline output (dict of DataFrames) to a single,
nicely formatted SSL_Dashboard.xlsx with auto-filters, frozen header
rows, conditional color scale on the SSL% columns, and column widths
sized to content.

Kept deliberately simple/openpyxl-only (no win32com), so it works the
same on Windows Task Scheduler or any other host (Linux/cron, etc.).
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import config

logger = logging.getLogger("ssl_dashboard.excel")

SSL_PCT_COLUMNS = {"SSL_QTY", "SSL_VALUE"}
SHEET_ORDER = [
    "Raw_Merged_Data",
    "SSL_Summary",
    "Vendor_Wise_View",
    "Brand_Wise_View",
    "Monthly_Wise_View",
    "Warehouse_Receipts",
]


def _write_sheet(ws: Worksheet, df: pd.DataFrame) -> None:
    header_fill = PatternFill("solid", fgColor=config.HEADER_FILL_COLOR)
    header_font = Font(color=config.HEADER_FONT_COLOR, bold=True)

    # Header row
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows — write via values for speed (avoid per-cell df.iat overhead
    # at 200k+ rows by using itertuples).
    pct_col_idx = {
        i + 1 for i, c in enumerate(df.columns) if c in SSL_PCT_COLUMNS
    }
    for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in pct_col_idx and value is not None:
                cell.number_format = '0.0"%"'

    # Freeze header + enable autofilter across full used range
    ws.freeze_panes = "A2"
    last_col_letter = get_column_letter(len(df.columns))
    last_row = max(len(df) + 1, 1)
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    # Conditional color scale on SSL% columns (red -> yellow -> green)
    for col_idx in pct_col_idx:
        col_letter = get_column_letter(col_idx)
        rng = f"{col_letter}2:{col_letter}{last_row}"
        rule = ColorScaleRule(
            start_type="num", start_value=0, start_color="F8696B",
            mid_type="num", mid_value=80, mid_color="FFEB84",
            end_type="num", end_value=100, end_color="63BE7B",
        )
        ws.conditional_formatting.add(rng, rule)

    # Column widths sized to header + a sample of content (cap to avoid
    # huge cost on a 200k-row sheet — sampling first 200 rows is enough
    # for a sane width estimate).
    sample = df.head(200)
    for col_idx, col_name in enumerate(df.columns, start=1):
        header_len = len(str(col_name))
        if len(sample) > 0:
            content_len = sample[col_name].astype(str).str.len().max()
        else:
            content_len = 0
        width = min(max(header_len, content_len) + 3, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_dashboard(sheets: dict[str, pd.DataFrame], output_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)  # drop default empty sheet

    for name in SHEET_ORDER:
        df = sheets.get(name)
        if df is None:
            logger.warning("Sheet '%s' missing from pipeline output — skipping.", name)
            continue
        ws = wb.create_sheet(title=name[:31])  # Excel sheet name limit
        logger.info("Writing sheet '%s' (%s rows x %s cols)", name, len(df), len(df.columns))
        _write_sheet(ws, df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Saved dashboard -> %s", output_path)
    return output_path


def save_with_history(sheets: dict[str, pd.DataFrame]) -> list[Path]:
    """Saves the main SSL_Dashboard.xlsx, plus an optional dated copy and
    an optional shared-folder copy, per config.py settings."""
    saved_paths = []

    main_path = config.OUTPUT_DIR / config.OUTPUT_FILENAME
    build_dashboard(sheets, main_path)
    saved_paths.append(main_path)

    if config.KEEP_DATED_COPY:
        dated_name = f"SSL_Dashboard_{datetime.now():%Y-%m-%d}.xlsx"
        dated_path = config.OUTPUT_DIR / "history" / dated_name
        dated_path.parent.mkdir(parents=True, exist_ok=True)
        build_dashboard(sheets, dated_path)
        saved_paths.append(dated_path)

    if config.SHARED_OUTPUT_DIR:
        shared_path = Path(config.SHARED_OUTPUT_DIR) / config.OUTPUT_FILENAME
        try:
            build_dashboard(sheets, shared_path)
            saved_paths.append(shared_path)
        except OSError as e:
            logger.error("Could not write to shared folder %s: %s", shared_path, e)

    return saved_paths
